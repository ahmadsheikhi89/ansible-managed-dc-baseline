#!/usr/bin/env python3
"""Build CTO-ready Excel workbook from synthetic baseline reports."""
from __future__ import annotations

import csv
import sys
from collections import Counter
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, PieChart, Reference
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError as exc:
    raise SystemExit(
        "Missing dependency: openpyxl. Install it from your internal mirror, for example: sudo dnf install -y python3-openpyxl"
    ) from exc

REPO_ROOT = Path(__file__).resolve().parents[1]
INPUTS = {
    "final": REPO_ROOT / "reports/final/99-managed-dc-baseline-readonly.csv",
    "lynis": REPO_ROOT / "reports/61-lynis-rocky-summary.csv",
    "exceptions": REPO_ROOT / "reports/exceptions/74-firewalld-operational-exceptions.csv",
    "firewall": REPO_ROOT / "reports/firewall/73-gitlab-firewalld-readonly-report.csv",
    "repo_cleanup": REPO_ROOT / "reports/repo-cleanup/69-external-repo-refs-report.csv",
}
OUTPUT = REPO_ROOT / "reports/final/managed-dc-baseline-cto.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="111827")
HEADER_FONT = Font(color="FFFFFF", bold=True)
SUBHEADER_FILL = PatternFill("solid", fgColor="E5E7EB")
OK_FILL = PatternFill("solid", fgColor="DCFCE7")
WARN_FILL = PatternFill("solid", fgColor="FEF3C7")
BORDER = Border(left=Side(style="thin", color="D1D5DB"), right=Side(style="thin", color="D1D5DB"), top=Side(style="thin", color="D1D5DB"), bottom=Side(style="thin", color="D1D5DB"))


def read_pipe_csv(path: Path) -> list[dict[str, str]]:
    if not path.exists():
        raise FileNotFoundError(f"Required input file not found: {path}")
    with path.open("r", newline="", encoding="utf-8") as handle:
        return list(csv.DictReader(handle, delimiter="|"))


def autosize(ws, max_width: int = 42) -> None:
    for column_cells in ws.columns:
        letter = get_column_letter(column_cells[0].column)
        width = 10
        for cell in column_cells:
            if cell.value is not None:
                width = max(width, min(max_width, len(str(cell.value)) + 2))
        ws.column_dimensions[letter].width = width


def style_table(ws, header_row: int = 1) -> None:
    ws.freeze_panes = f"A{header_row + 1}"
    for cell in ws[header_row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    for row in ws.iter_rows(min_row=header_row + 1):
        for cell in row:
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if str(cell.value).upper() in {"OK", "YES", "ACTIVE", "ENFORCING"}:
                cell.fill = OK_FILL
            if str(cell.value).upper() in {"ACCEPTED_EXCEPTION", "UBUNTU_EXCEPTION", "INACTIVE"}:
                cell.fill = WARN_FILL
    autosize(ws)


def write_rows(ws, rows: list[dict[str, str]]) -> None:
    if not rows:
        return
    headers = list(rows[0].keys())
    ws.append(headers)
    for row in rows:
        ws.append([row.get(h, "") for h in headers])
    style_table(ws)


def add_kpi_block(ws, final_rows: list[dict[str, str]]) -> None:
    os_counts = Counter(row["os"] for row in final_rows)
    fw_counts = Counter(row["firewall_control"] for row in final_rows)
    final_counts = Counter(row["final_status"] for row in final_rows)
    kpis = [
        ("Total Hosts", len(final_rows)),
        ("Final OK", final_counts.get("OK", 0)),
        ("Rocky Linux", os_counts.get("Rocky", 0)),
        ("Ubuntu Exception", os_counts.get("Ubuntu", 0)),
        ("External Repo Refs", sum(int(row["external_repo_refs"]) for row in final_rows)),
        ("Docker Active", sum(1 for row in final_rows if row["docker"] == "active")),
        ("Time Sync Yes", sum(1 for row in final_rows if row["time_sync"] == "yes")),
        ("SELinux Enforcing", sum(1 for row in final_rows if row["selinux"] == "Enforcing")),
    ]
    ws["A1"] = "AtlasForge Bank - Managed DC Baseline Dashboard"
    ws["A1"].font = Font(size=16, bold=True, color="111827")
    ws["A2"] = "Synthetic public-safe executive report generated from local CSV evidence."
    ws["A2"].font = Font(italic=True, color="4B5563")
    start_row = 4
    ws.cell(start_row, 1, "KPI")
    ws.cell(start_row, 2, "Value")
    for cell in ws[start_row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = BORDER
    for idx, (label, value) in enumerate(kpis, start_row + 1):
        ws.cell(idx, 1, label)
        ws.cell(idx, 2, value)
        ws.cell(idx, 1).border = BORDER
        ws.cell(idx, 2).border = BORDER
        ws.cell(idx, 2).alignment = Alignment(horizontal="center")
        ws.cell(idx, 2).fill = OK_FILL if label not in {"External Repo Refs"} or value == 0 else WARN_FILL

    sections = {
        "OS Distribution": os_counts,
        "Final Status": final_counts,
        "Firewall Control": fw_counts,
        "Control Readiness": Counter({
            "Hosts Entry Ready": sum(1 for row in final_rows if row["hub_hosts_entry"] == "yes" and row["git_hosts_entry"] == "yes"),
            "Repo Clean": sum(1 for row in final_rows if row["external_repo_refs"] == "0"),
            "Time Sync Ready": sum(1 for row in final_rows if row["time_sync"] == "yes"),
            "Docker Ready": sum(1 for row in final_rows if row["docker"] == "active"),
        }),
    }
    col = 4
    chart_positions = [("G4", "M18"), ("G20", "M34"), ("N4", "T18"), ("N20", "T34")]
    for chart_index, (title, counter) in enumerate(sections.items()):
        table_col = col + (chart_index % 2) * 4
        table_row = 4 + (chart_index // 2) * 16
        ws.cell(table_row, table_col, title)
        ws.cell(table_row, table_col + 1, "Count")
        for c in [ws.cell(table_row, table_col), ws.cell(table_row, table_col + 1)]:
            c.fill = HEADER_FILL
            c.font = HEADER_FONT
            c.border = BORDER
        for offset, (label, value) in enumerate(counter.items(), 1):
            ws.cell(table_row + offset, table_col, label)
            ws.cell(table_row + offset, table_col + 1, value)
            ws.cell(table_row + offset, table_col).border = BORDER
            ws.cell(table_row + offset, table_col + 1).border = BORDER
        data = Reference(ws, min_col=table_col + 1, min_row=table_row, max_row=table_row + len(counter))
        labels = Reference(ws, min_col=table_col, min_row=table_row + 1, max_row=table_row + len(counter))
        if title in {"OS Distribution", "Final Status"}:
            chart = PieChart()
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(labels)
            chart.title = title
        else:
            chart = BarChart()
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(labels)
            chart.title = title
            chart.y_axis.title = "Count"
            chart.x_axis.title = "Control"
            chart.height = 7
            chart.width = 12
        ws.add_chart(chart, chart_positions[chart_index][0])
    autosize(ws)


def build_yes_no_matrix(final_rows: list[dict[str, str]]) -> list[dict[str, str]]:
    matrix = []
    for row in final_rows:
        matrix.append({
            "inventory_hostname": row["inventory_hostname"],
            "hosts_entries": "YES" if row["hub_hosts_entry"] == "yes" and row["git_hosts_entry"] == "yes" else "NO",
            "repo_clean": "YES" if row["external_repo_refs"] == "0" else "NO",
            "time_sync": "YES" if row["time_sync"] == "yes" else "NO",
            "docker_active": "YES" if row["docker"] == "active" else "NO",
            "selinux_ready": "YES" if row["selinux"] in {"Enforcing", "NotApplicable"} else "NO",
            "firewall_ready": "YES" if row["firewall_control"] in {"OK", "ACCEPTED_EXCEPTION", "UBUNTU_EXCEPTION"} else "NO",
            "final_status": row["final_status"],
        })
    return matrix


def build_control_checklist(final_rows: list[dict[str, str]]) -> list[dict[str, str]]:
    total = len(final_rows)
    return [
        {"control_id": "CTRL-001", "control_name": "Inventory scope documented", "ready_hosts": str(total), "total_hosts": str(total), "status": "OK", "evidence": "inventory/hosts.yml"},
        {"control_id": "CTRL-002", "control_name": "Internal hosts entries validated", "ready_hosts": str(sum(1 for r in final_rows if r["hub_hosts_entry"] == "yes" and r["git_hosts_entry"] == "yes")), "total_hosts": str(total), "status": "OK", "evidence": "reports/final/99-managed-dc-baseline-readonly.csv"},
        {"control_id": "CTRL-003", "control_name": "External repo references removed", "ready_hosts": str(sum(1 for r in final_rows if r["external_repo_refs"] == "0")), "total_hosts": str(total), "status": "OK", "evidence": "reports/repo-cleanup/69-external-repo-refs-report.csv"},
        {"control_id": "CTRL-004", "control_name": "Time synchronization validated", "ready_hosts": str(sum(1 for r in final_rows if r["time_sync"] == "yes")), "total_hosts": str(total), "status": "OK", "evidence": "reports/final/99-managed-dc-baseline-readonly.csv"},
        {"control_id": "CTRL-005", "control_name": "Docker runtime active", "ready_hosts": str(sum(1 for r in final_rows if r["docker"] == "active")), "total_hosts": str(total), "status": "OK", "evidence": "reports/final/99-managed-dc-baseline-readonly.csv"},
        {"control_id": "CTRL-006", "control_name": "Rocky SELinux enforcing", "ready_hosts": str(sum(1 for r in final_rows if r["selinux"] == "Enforcing")), "total_hosts": "19", "status": "OK", "evidence": "reports/final/99-managed-dc-baseline-readonly.csv"},
        {"control_id": "CTRL-007", "control_name": "Firewall exceptions documented", "ready_hosts": "2", "total_hosts": "2", "status": "OK", "evidence": "reports/exceptions/74-firewalld-operational-exceptions.csv"},
    ]


def main() -> None:
    try:
        data = {key: read_pipe_csv(path) for key, path in INPUTS.items()}
    except FileNotFoundError as exc:
        print(str(exc), file=sys.stderr)
        print("Run python3 scripts/generate-demo-data.py first.", file=sys.stderr)
        raise SystemExit(1)

    wb = Workbook()
    default = wb.active
    wb.remove(default)

    dashboard = wb.create_sheet("00_Dashboard")
    add_kpi_block(dashboard, data["final"])

    ws = wb.create_sheet("01_Final_Baseline")
    write_rows(ws, data["final"])

    ws = wb.create_sheet("02_Yes_No_Matrix")
    write_rows(ws, build_yes_no_matrix(data["final"]))

    ws = wb.create_sheet("03_CIS_Lynis")
    write_rows(ws, data["lynis"])

    ws = wb.create_sheet("04_Exceptions")
    write_rows(ws, data["exceptions"])

    ws = wb.create_sheet("05_Firewall_Evidence")
    write_rows(ws, data["firewall"])

    ws = wb.create_sheet("06_Repo_Cleanup")
    write_rows(ws, data["repo_cleanup"])

    ws = wb.create_sheet("07_Control_Checklist")
    write_rows(ws, build_control_checklist(data["final"]))

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT)
    print(f"Excel workbook generated: {OUTPUT}")


if __name__ == "__main__":
    main()
